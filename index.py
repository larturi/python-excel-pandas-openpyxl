import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

def automatizar_excel(file_name):
    # Leo el archivo excel
    excel_file = pd.read_excel(file_name)

    # Creo una lista con los nombres de las columnas que necesito
    filtered_data = excel_file[['Gender', 'Product line', 'Total']]

    # Creo una tabla pivote
    tabla_pivote = filtered_data.pivot_table(
        index='Gender', 
        columns='Product line', 
        values='Total', 
        aggfunc='sum').round(0)

    # Exporto a excel sales_gender.xlsx
    tabla_pivote.to_excel('report_sales_gender.xlsx', startrow=4, sheet_name='Sales Gender')

    # Abro el archivo sales_gender.xlsx
    wb = load_workbook('report_sales_gender.xlsx')
    sheet = wb['Sales Gender']

    # Limites de sales_gender.xlsx
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Grafico
    barchart = BarChart()
    data = Reference(sheet, min_col=min_column+1, min_row=min_row, max_col=max_column, max_row=max_row)
    categories = Reference(sheet, min_col=min_column, min_row=min_row + 1, max_col=min_column, max_row=max_row)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, 'A10')
    barchart.title = 'Ventas'
    barchart.style = 2

    # Totales
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] #note: Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    # Sum in columns B-G
    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'

    # Formatting the report
    sheet['A1'] = 'Sales Report'
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
        
    wb.save('sales_gender.xlsx')
    
    return

automatizar_excel('supermarket_sales.xlsx')
